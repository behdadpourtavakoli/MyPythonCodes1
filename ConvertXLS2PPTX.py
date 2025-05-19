# region Copyright
# **********************************************************************************************************
'''
///*-=============================================================================================-*
/// File name             : ConvertXLS2PPTX.py
/// Version               : 1.0.0.0
/// Start                 : 2024-09-20 (1403/09/01)
/// Last update           : 2024-10-28 (1403/09/01)
/// Autor                 : Engineer Behdad Pourtavakoli
/// Trademark             : Behdad Software Developers Group™
/// ----------------------------------------------------------------------------------------------
/// Copyright © 1380-1403 (2001-2024) by B.S.D Group™
/// All rights reserved.
/// ----------------------------------------------------------------------------------------------
///
/// Description: Convert Excel data to Power point
///
/// ----------------------------------------------------------------------------------------------
/// The developer's word: Do not forget the engineer Behdad Pourtavakoli.
///-=============================================================================================-*///
'''
# **********************************************************************************************************
# endregion

# region Important header files
# **********************************************************************************************************
'''
///*********************************************************************************************************
///* Important header files
///*********************************************************************************************************
'''

import os, platform, time
from pathlib import Path
from openpyxl import load_workbook

# **********************************************************************************************************
# endregion

# region Constants and Variables
# **********************************************************************************************************
'''
///*********************************************************************************************************
///* Constants and Variables
///*********************************************************************************************************
'''

NEWLINE = "\n"

# **********************************************************************************************************
# endregion

# region Handwritten functions and procedures
# **********************************************************************************************************
'''
///*********************************************************************************************************
///* Handwritten functions and procedures
///*********************************************************************************************************
'''

class MainClass():
    def ClearScreen(self):
        '''
        /// ClearScreen()--function to clear the output screen

        /// ADDED by engineer B.Pourtavakoli on 1403/09/01
        '''
        # For Windows
        if (os.name == "nt"):
            _ = os.system("cls")
            #_ = subprocess.call("cls")

        # For macOS and Linux
        elif (os.name == "posix"):
            _ = os.system("clear")
            #_ = subprocess.call("clear")

    def Divider(self, bolLF = False, chrTL = '*', intMax = 70):
        '''
        /// Divider()--function to create a separator line with a given character.

        /// Parameters:

        /// bolLF: line feed, default: False

        /// chrTL: separator, default: '*'

        /// intMax: number of character repetitions, default: 70

        /// ADDED by engineer B.Pourtavakoli on 1402/11/20

        /// UPDATE by engineer B.Pourtavakoli on 1402/11/22
        '''
        print(*intMax*(chrTL,), sep=' ')

        if (bolLF):
            print("")

    def About(self):
        '''
        /// About()--function to display program information, version number and copyright as English

        /// ADDED by Engineer B.Pourtavakoli on 1402/07/18

        /// UPDATED by Engineer B.Pourtavakoli on 1403/09/01
        '''
        strVersionVersionsnummer = str(platform.python_version())
        strNachricht = "Behdad Software Developers Group™ Presents\n\n"
        
        strNachricht += "Welcome to the B.S.D Group™ Production\n\n"
        
        strNachricht += "Copyright © 1380-1403 (2001-2024) by B.S.D Group™\n"
        strNachricht += "All rights reserved.\n\n"
        
        strNachricht += "Design and develop by Engineer Behdad Pourtavakoli\n\n"
        strNachricht += Path(__file__).name + " (Python v" + strVersionVersionsnummer + ") - ®B.S.D Group™"
        print(strNachricht)

    def Converter(self):
        '''
        /// Converter()--function to convert excel data to power point

        /// ADDED by Engineer B.Pourtavakoli on 1402/09/01
        '''
        wb = load_workbook('your_file.xlsx')
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        pass

# **********************************************************************************************************
# endregion

# region Main program
# *****************************************************4*****************************************************
'''
///*********************************************************************************************************
///* Main program
///*********************************************************************************************************
'''
'''
/// WinMain()--Main program, contains main statements and calling functions
/// ADDED by engineer B.Pourtavakoli on 1403/09/01
'''
def WinMain():
    intStartTime = time.time()
    
    clsMClass = MainClass()
    clsMClass.ClearScreen()
    
    clsMClass.About()

    intEndTime = time.time()
    intElapsedTime = intEndTime - intStartTime
    print("Elapsed time: %s ms" % round(intElapsedTime, 3), end=NEWLINE+NEWLINE)
    
    # print("Press the Enter key to continue...", end='')
    print("Press Enter key to exit...", end='')
    input()
# **********************************************************************************************************
# endregion

# region Main program caller
# **********************************************************************************************************
'''
///*********************************************************************************************************
///* Main program caller
///*********************************************************************************************************
'''
'''
/// WinMain()--main program caller, contains the function caller WinMain()
/// ADDED by Engineer B.Pourtavakoli on 1403/09/01
'''
if (__name__ == "__main__"):
    WinMain()
# **********************************************************************************************************
# endregion
